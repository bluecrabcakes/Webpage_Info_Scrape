from bs4 import BeautifulSoup
import requests
import openpyxl as xl

wb = xl.load_workbook("Book3.xlsx")
sheet = wb['Sheet1']

for row in range(1207, 1407):
    source = requests.get(sheet.cell(row, 1).value).text
    soup = BeautifulSoup(source, 'lxml')

    name = soup.find('h3').text
    mobile = soup.find(id="ContentPlaceHolder1_DataList3_Label9_0").text
    phone = soup.find(id="ContentPlaceHolder1_DataList3_Label8_0").text
    email = soup.find(id="ContentPlaceHolder1_DataList3_Label10_0").text
    website = soup.find(id="ContentPlaceHolder1_DataList3_Label11_0").text

    name_cell = sheet.cell(row, 2)
    tele_cell = sheet.cell(row, 3)
    email_cell = sheet.cell(row, 4)
    web_cell = sheet.cell(row, 5)

    name_cell.value = name
    tele_cell.value = mobile + '     ' + phone
    email_cell.value = email
    web_cell.value = website

wb.save("Book3.xlsx")
